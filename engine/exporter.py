"""
Leg summary exporter.

Produces, for each closed (or manually triggered) leg:
  • One .xlsx workbook with sheets: Summary, Systems, Bodies,
    Terraformable, Body Classes, Organics, Codex Firsts, Sales.
  • Flat .csv files for each detail table.
  • A master rollup workbook aggregating all legs.

Metric definitions (see db.py for details):
  • Bodies scanned / first-discovered / surface-mapped
      → DISTINCT (system, body_id) rows from the bodies table.
        Duplicate scan events for revisited systems do not inflate these.
  • Terraformable bodies
      → distinct bodies where TerraformState is non-empty, regardless of class.
        This is the primary high-value exploration metric.
  • Body-class breakdown
      → per-class counts from the bodies table, reported independently.
        Stars (spectral codes) and planets (full names) both appear.
  • Organic variants → distinct Variant_Localised strings at stage 3.
"""
from __future__ import annotations

import csv
import json
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .db import Database

log = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")

# Classes always shown in the Notable/Terraformable sheet even without terraform state
_HIGH_VALUE_CLASSES = {"Earthlike body", "Water world", "Ammonia world"}

# Star spectral type codes — used to separate stars from planets in class breakdown
_STAR_TYPES = {
    "O", "B", "A", "F", "G", "K", "M", "L", "T", "Y",
    "TTS", "AeBe", "W", "WN", "WNC", "WC", "WO",
    "CS", "C", "CN", "CJ", "CH", "CHd", "MS", "S",
    "D", "DA", "DAB", "DAO", "DAZ", "DAV",
    "DB", "DBZ", "DBV", "DO", "DOV", "DQ", "DC", "DCV", "DX",
    "N", "H", "SupermassiveBlackHole",
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws) -> None:
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _write_csv(path: Path, headers: list[str], rows: list[list]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        w.writerows(rows)


def _leg_slug(leg) -> str:
    name = (leg["name"] or f"Leg{leg['ordinal']:02d}").replace(" ", "_")
    date = (leg["start_ts"] or "")[:10]
    return f"Leg{leg['ordinal']:02d}_{name}_{date}"


def _add_sheet(wb, name: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    _style_header(ws)
    for i, row in enumerate(rows, 2):
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = _ALT_FILL
    _autofit(ws)


# ── Main entry points ──────────────────────────────────────────────────────────

def export_leg(leg_id: int, db: Database, output_dir: Path,
               suffix: str = "") -> Path:
    """Export one leg to xlsx + CSVs.  Returns the xlsx path.

    Pass suffix (e.g. '_snapshot_20260603T014500') to distinguish an
    on-demand snapshot from the final waypoint-close export.
    """
    leg = db.get_leg(leg_id)
    if leg is None:
        raise ValueError(f"Leg {leg_id} not found in DB.")

    slug   = _leg_slug(leg) + suffix
    xlsx_p = output_dir / f"{slug}.xlsx"
    wb     = openpyxl.Workbook()
    wb.remove(wb.active)

    stats     = db.stats_for_leg(leg_id)          # DISTINCT counts
    class_bkd = db.get_body_class_breakdown(leg_id)
    rf_stats  = db.get_rare_finds_stats(leg_id)

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    summary_rows: list[tuple] = [
        ("Leg",                  leg["ordinal"]),
        ("Name",                 leg["name"] or "—"),
        ("Start system",         leg["start_system"] or "—"),
        ("Start time (UTC)",     leg["start_ts"] or "—"),
        ("End system",           leg["end_system"] or "—"),
        ("End time (UTC)",       leg["end_ts"] or "—"),
        ("", ""),
        ("— Journey ——————————", ""),
        ("Commander jumps",      stats["commander_jumps"]),
        ("Total distance (LY)",  f"{stats['distance_ly']:.2f}"),
        ("Systems visited",      stats["systems_visited"]),
        ("", ""),
        ("— Discoveries (distinct bodies) ——", ""),
        ("Bodies scanned",       stats["bodies_scanned"]),
        ("First-discovered",     stats["first_discovered"]),
        ("Surface-mapped",       stats["bodies_mapped"]),
        ("Terraformable bodies", stats["terraformable_count"]),
        ("", ""),
        ("— Biology / Codex ——", ""),
        ("Distinct organic variants", stats["organic_variants"]),
        ("New codex entries",    stats["new_codex"]),
        ("", ""),
        ("— Sales ————————————", ""),
        ("Cartographics sales",  stats["carto_sales_count"]),
        ("Cartographics cr",     stats["carto_earnings"]),
        ("Exo-Bio sales",        stats["exobio_sales_count"]),
        ("Exo-Bio cr",           stats["exobio_earnings"]),
        ("", ""),
        ("— High-value bodies ————————————", ""),
        ("Earth-likes",          class_bkd.get("Earthlike body", 0)),
        ("Water worlds",         class_bkd.get("Water world", 0)),
        ("Ammonia worlds",       class_bkd.get("Ammonia world", 0)),
        ("High-metal-content",   class_bkd.get("High metal content body", 0)),
        ("", ""),
        ("— Rare / notable finds ————————", ""),
        ("Rare finds (first discoveries)", rf_stats["rare_finds_count"]),
    ]
    rf_bkd = rf_stats["rare_finds_by_rule"]
    for rule_tag, cnt in sorted(rf_bkd.items()):
        summary_rows.append((f"  {rule_tag}", cnt))
    summary_rows.append(("", ""))
    for r, (k, v) in enumerate(summary_rows, 1):
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22

    # ── Systems sheet ──────────────────────────────────────────────────────────
    sys_headers = [
        "System", "First Visit (UTC)", "Body Count", "Fully Scanned",
        "Star Pos X", "Star Pos Y", "Star Pos Z",
    ]
    sys_rows = []
    for s in db.get_systems_for_leg(leg_id):
        sys_rows.append([
            s["system_name"], s["first_visit_ts"], s["body_count"],
            "Yes" if s["fully_scanned"] else "No",
            s["star_pos_x"], s["star_pos_y"], s["star_pos_z"],
        ])
    _add_sheet(wb, "Systems", sys_headers, sys_rows)
    _write_csv(output_dir / f"{slug}_systems.csv", sys_headers, sys_rows)

    # ── All bodies sheet ───────────────────────────────────────────────────────
    body_headers = [
        "System", "Body Name", "Body ID", "Class", "Terraform State",
        "Landable", "Dist LS", "First Discovery", "Pre-Mapped", "Surface Mapped",
    ]
    body_rows = []
    for b in db.get_bodies_for_leg(leg_id):
        body_rows.append([
            b["system"], b["body_name"], b["body_id"], b["body_class"],
            b["terraform_state"],
            "Yes" if b["landable"] else "No",
            b["distance_from_arrival_ls"],
            "Yes" if not b["was_discovered"] else "No",
            "Yes" if not b["was_mapped"] else "No",
            "Yes" if b["surface_mapped"] else "No",
        ])
    _add_sheet(wb, "Bodies", body_headers, body_rows)
    _write_csv(output_dir / f"{slug}_bodies.csv", body_headers, body_rows)

    # ── Terraformable bodies sheet ─────────────────────────────────────────────
    # Includes ALL terraformable bodies (any class) plus the high-value non-
    # terraformable classes (ELW, WW, AW) that are always worth calling out.
    terra_headers = [
        "System", "Body Name", "Class", "Terraform State",
        "First Discovery", "Surface Mapped", "Landable",
    ]
    terra_rows = []
    for b in db.get_bodies_for_leg(leg_id):
        is_terra = bool(b["terraform_state"] and b["terraform_state"].strip())
        is_high_value = b["body_class"] in _HIGH_VALUE_CLASSES
        if is_terra or is_high_value:
            terra_rows.append([
                b["system"], b["body_name"], b["body_class"],
                b["terraform_state"] or "",
                "Yes" if not b["was_discovered"] else "No",
                "Yes" if b["surface_mapped"] else "No",
                "Yes" if b["landable"] else "No",
            ])
    _add_sheet(wb, "Terraformable & Notable", terra_headers, terra_rows)
    _write_csv(output_dir / f"{slug}_terraformable.csv", terra_headers, terra_rows)

    # ── Body-class breakdown sheet ─────────────────────────────────────────────
    # Separate stars (spectral codes) from planets (full English names).
    bkd_headers = ["Class", "Count", "Type"]
    bkd_rows = []
    for cls, cnt in class_bkd.items():
        kind = "Star" if cls in _STAR_TYPES else "Planet/Other"
        bkd_rows.append([cls, cnt, kind])
    _add_sheet(wb, "Body Classes", bkd_headers, bkd_rows)
    _write_csv(output_dir / f"{slug}_body_classes.csv", bkd_headers, bkd_rows)

    # ── Organics sheet ─────────────────────────────────────────────────────────
    org_headers = ["System", "Body ID", "Genus", "Species", "Variant"]
    org_rows = []
    for o in db.get_organics_for_leg(leg_id):
        org_rows.append([o["system"], o["body_id"], o["genus"], o["species"], o["variant"]])
    _add_sheet(wb, "Organics", org_headers, org_rows)
    _write_csv(output_dir / f"{slug}_organics.csv", org_headers, org_rows)

    # ── Codex sheet ────────────────────────────────────────────────────────────
    codex_headers = ["Timestamp", "Name", "Region", "System"]
    codex_rows = [[c["ts"], c["name"], c["region"], c["system"]]
                  for c in db.get_codex_for_leg(leg_id)]
    _add_sheet(wb, "Codex Firsts", codex_headers, codex_rows)
    _write_csv(output_dir / f"{slug}_codex.csv", codex_headers, codex_rows)

    # ── Rare / notable finds sheet ─────────────────────────────────────────────
    rf_headers = [
        "System", "Body", "Class", "Rules Matched", "Trigger Details",
        "Dist LS", "First Discovery", "Pre-Mapped", "Footfalled",
    ]
    rf_rows = []
    for rf in db.get_rare_finds_for_leg(leg_id):
        try:
            rules = ", ".join(json.loads(rf["matched_rules"]))
            dets  = "; ".join(
                f"{k}: {v}" for k, v in json.loads(rf["trigger_details"]).items()
            )
        except Exception:
            rules = rf["matched_rules"]
            dets  = rf["trigger_details"]
        rf_rows.append([
            rf["system"],
            rf["body_name"] or "",
            rf["body_class"] or "",
            rules,
            dets,
            rf["distance_ls"],
            "Yes" if not rf["was_discovered"] else "No",
            "Yes" if not rf["was_mapped"]     else "No",
            "Yes" if not rf["was_footfalled"] else "No",
        ])
    _add_sheet(wb, "Rare Finds", rf_headers, rf_rows)
    _write_csv(output_dir / f"{slug}_rare_finds.csv", rf_headers, rf_rows)

    # ── Sales sheet ────────────────────────────────────────────────────────────
    sales_headers = ["Type", "Timestamp", "Total Earnings", "Base Value", "Bonus"]
    sales_rows = []
    for s in db.get_carto_sales_for_leg(leg_id):
        sales_rows.append(["Cartographics", s["ts"], s["total_earnings"], s["base_value"], s["bonus"]])
    for s in db.get_organic_sales_for_leg(leg_id):
        sales_rows.append(["Exo-Bio", s["ts"], s["value"] + s["bonus"], s["value"], s["bonus"]])
    _add_sheet(wb, "Sales", sales_headers, sales_rows)
    _write_csv(output_dir / f"{slug}_sales.csv", sales_headers, sales_rows)

    wb.save(str(xlsx_p))
    log.info("Exported leg %d -> %s", leg_id, xlsx_p.name)
    return xlsx_p


def export_master_rollup(db: Database, output_dir: Path) -> Path:
    """Build/overwrite the master rollup workbook."""
    xlsx_p = output_dir / "master_rollup.xlsx"
    wb     = openpyxl.Workbook()
    wb.remove(wb.active)

    all_legs = db.get_all_legs()

    # ── Per-leg summary table ──────────────────────────────────────────────────
    ws = wb.create_sheet("Legs")
    headers = [
        "Leg", "Name", "Start System", "Start UTC", "End System", "End UTC",
        "Status", "Jumps", "Dist LY", "Systems",
        "Bodies", "1st Disc", "Mapped", "Terraformable",
        "Organics", "Codex", "Carto Cr", "ExoBio Cr", "Rare Finds",
    ]
    ws.append(headers)
    _style_header(ws)

    grand     = db.stats_for_leg(None)
    grand_bkd = db.get_body_class_breakdown(None)
    grand_rf  = db.get_rare_finds_stats(None)

    for leg in all_legs:
        s  = db.stats_for_leg(leg["leg_id"])
        rf = db.get_rare_finds_stats(leg["leg_id"])
        ws.append([
            leg["ordinal"],
            leg["name"] or "—",
            leg["start_system"] or "—",
            leg["start_ts"] or "—",
            leg["end_system"] or "—",
            leg["end_ts"] or "—",
            leg["status"],
            s["commander_jumps"],
            s["distance_ly"],
            s["systems_visited"],
            s["bodies_scanned"],
            s["first_discovered"],
            s["bodies_mapped"],
            s["terraformable_count"],
            s["organic_variants"],
            s["new_codex"],
            s["carto_earnings"],
            s["exobio_earnings"],
            rf["rare_finds_count"],
        ])

    ws.append([])
    ws.append([
        "TOTAL", "", "", "", "", "", "",
        grand["commander_jumps"], grand["distance_ly"], grand["systems_visited"],
        grand["bodies_scanned"], grand["first_discovered"],
        grand["bodies_mapped"], grand["terraformable_count"],
        grand["organic_variants"], grand["new_codex"],
        grand["carto_earnings"], grand["exobio_earnings"],
        grand_rf["rare_finds_count"],
    ])
    _autofit(ws)

    # ── Whole-expedition summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Expedition Summary")
    summ = [
        ("Commander jumps",     grand["commander_jumps"]),
        ("Total distance (LY)", f"{grand['distance_ly']:.2f}"),
        ("Carrier jumps",       grand["carrier_jumps"]),
        ("Systems visited",     grand["systems_visited"]),
        ("Bodies scanned",      grand["bodies_scanned"]),
        ("First-discovered",    grand["first_discovered"]),
        ("Surface-mapped",      grand["bodies_mapped"]),
        ("Terraformable",       grand["terraformable_count"]),
        ("Distinct organics",   grand["organic_variants"]),
        ("New codex entries",   grand["new_codex"]),
        ("Earth-likes",         grand_bkd.get("Earthlike body", 0)),
        ("Water worlds",        grand_bkd.get("Water world", 0)),
        ("Ammonia worlds",      grand_bkd.get("Ammonia world", 0)),
        ("High-metal-content",  grand_bkd.get("High metal content body", 0)),
        ("Carto sales",         grand["carto_sales_count"]),
        ("Carto earnings",      grand["carto_earnings"]),
        ("Exo-Bio sales",       grand["exobio_sales_count"]),
        ("Exo-Bio earnings",    grand["exobio_earnings"]),
        ("", ""),
        ("— Rare / notable (first discoveries) ——", ""),
        ("Total rare finds",    grand_rf["rare_finds_count"]),
    ]
    for rule_tag, cnt in sorted(grand_rf["rare_finds_by_rule"].items()):
        summ.append((f"  {rule_tag}", cnt))
    for r, (k, v) in enumerate(summ, 1):
        ws2.cell(r, 1, k)
        ws2.cell(r, 2, v)
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 20

    # ── Body-class breakdown for whole expedition ──────────────────────────────
    ws3 = wb.create_sheet("Body Classes")
    ws3.append(["Class", "Count", "Type"])
    _style_header(ws3)
    for cls, cnt in grand_bkd.items():
        kind = "Star" if cls in _STAR_TYPES else "Planet/Other"
        ws3.append([cls, cnt, kind])
    _autofit(ws3)

    # ── Rare finds breakdown for whole expedition ──────────────────────────────
    ws4 = wb.create_sheet("Rare Finds")
    ws4.append(["Rule", "Count"])
    _style_header(ws4)
    for rule_tag, cnt in sorted(grand_rf["rare_finds_by_rule"].items()):
        ws4.append([rule_tag, cnt])
    ws4.append([])
    ws4.append(["TOTAL", grand_rf["rare_finds_count"]])
    _autofit(ws4)

    # ── All-expedition rare finds detail ──────────────────────────────────────
    all_rf_headers = [
        "Leg", "System", "Body", "Class", "Rules Matched", "Trigger Details",
        "Dist LS", "First Discovery", "Pre-Mapped", "Footfalled",
    ]
    all_rf_rows = []
    for leg in all_legs:
        for rf in db.get_rare_finds_for_leg(leg["leg_id"]):
            try:
                rules = ", ".join(json.loads(rf["matched_rules"]))
                dets  = "; ".join(
                    f"{k}: {v}" for k, v in json.loads(rf["trigger_details"]).items()
                )
            except Exception:
                rules = rf["matched_rules"]
                dets  = rf["trigger_details"]
            all_rf_rows.append([
                leg["ordinal"],
                rf["system"],
                rf["body_name"] or "",
                rf["body_class"] or "",
                rules,
                dets,
                rf["distance_ls"],
                "Yes" if not rf["was_discovered"] else "No",
                "Yes" if not rf["was_mapped"]     else "No",
                "Yes" if not rf["was_footfalled"] else "No",
            ])
    _add_sheet(wb, "Rare Finds Detail", all_rf_headers, all_rf_rows)

    wb.save(str(xlsx_p))
    log.info("Master rollup written -> %s", xlsx_p.name)
    return xlsx_p
